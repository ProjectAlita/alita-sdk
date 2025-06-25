import json
import pandas as pd
from urllib.parse import urlparse
from collections import OrderedDict
import re
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# Path to the Lighthouse JSON report
report_path = 'user-flow.report.json'

# Audits to extract from the report
performance_audits = [
    'first-contentful-paint',
    'speed-index',
    'interactive',
    'total-blocking-time',
    'largest-contentful-paint',
    'cumulative-layout-shift',
    'network-requests',
    'bootup-time',
    'interaction-to-next-paint',
    'server-response-time',
]

# Create lists for audits to be appended with ", sec" and ", ms"
sec_audits = ['first-contentful-paint', 'interactive', 'largest-contentful-paint', 'mainthread-work-breakdown', 'network-requests', 'speed-index', 'javaScript-execution-time']
ms_audits = ['interaction-to-next-paint', 'total-blocking-time', 'time-to-first-byte']

# Rename mappings
rename_audits = {
    'bootup-time': 'javaScript-execution-time',
    'server-response-time': 'time-to-first-byte'
}

def extract_application_name(url):
    parsed_url = urlparse(url)
    hostname_parts = parsed_url.hostname.split('.') if parsed_url.hostname else []
    application_name = hostname_parts[0] if len(hostname_parts) > 1 else '3rd-party'
    return application_name

def apply_excel_conditional_formatting(ws, column_letter, thresholds, colors):
    ws.conditional_formatting.add(f'{column_letter}2:{column_letter}{ws.max_row}',
                                  CellIsRule(operator='lessThanOrEqual', formula=[str(thresholds[0])], stopIfTrue=True, fill=PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")))
    ws.conditional_formatting.add(f'{column_letter}2:{column_letter}{ws.max_row}',
                                  CellIsRule(operator='between', formula=[str(thresholds[0]+0.0001), str(thresholds[1])], stopIfTrue=True, fill=PatternFill(start_color=colors[1], end_color=colors[1], fill_type="solid")))
    ws.conditional_formatting.add(f'{column_letter}2:{column_letter}{ws.max_row}',
                                  CellIsRule(operator='greaterThanOrEqual', formula=[str(thresholds[1]+0.0001)], stopIfTrue=True, fill=PatternFill(start_color=colors[2], end_color=colors[2], fill_type="solid")))

def apply_all_borders(ws):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

try:
    with open(report_path, 'r', encoding='utf-8') as file:
        data = json.load(file)

    data_rows = []
    step_order = OrderedDict()

    for index, step in enumerate(data.get('steps', [])):
        step_name = step.get('name', 'unknown_step')
        step_order[step_name] = index
        lhr_data = step.get('lhr', {})
        url = lhr_data.get('finalDisplayedUrl', '3rd-party')
        application_name = extract_application_name(url)

        performance_score = lhr_data.get('categories', {}).get('performance', {}).get('score')
        performance_score = performance_score * 100 if performance_score is not None else None

        for audit in performance_audits:
            audit_result = lhr_data.get('audits', {}).get(audit, {})
            numeric_value = audit_result.get('displayValue')

            if numeric_value is not None:
                numeric_value = re.sub(r'[a-zA-Z\s]', '', numeric_value)
                if numeric_value:
                    numeric_value = numeric_value.replace(',', '')
                    numeric_value = float(numeric_value)
                else:
                    numeric_value = None

            audit_display_name = rename_audits.get(audit, audit)
            if audit_display_name in sec_audits:
                audit_display_name += ", sec"
            elif audit_display_name in ms_audits:
                audit_display_name += ", ms"

            data_row = {"Step name": step_name,
                        "Performance Score": performance_score,
                        "Audit": audit_display_name,
                        "Numeric Value": numeric_value
                        }
            data_rows.append(data_row)

    df = pd.DataFrame(data_rows)
    df = df.pivot_table(index="Step name", columns="Audit", values="Numeric Value", aggfunc='mean')
    df = df.fillna('')
    df = df.reindex(step_order.keys())

    # Save DataFrame to Excel using openpyxl for styling
    writer = pd.ExcelWriter("output.xlsx", engine='openpyxl')
    df.to_excel(writer, index=True)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Apply styles
    header_fill = PatternFill(start_color="7FD5D8", end_color="7FD5D8", fill_type="solid")
    for cell in worksheet[1]:  # Apply styles to header row
        cell.fill = header_fill

    # Set alignment for 'Step name' column
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')

    # Apply conditional formatting
    for col_index, col_name in enumerate(df.columns, 2):  # Start from 2 to account for index column
        column_letter = get_column_letter(col_index)
        if col_name in ["cumulative-layout-shift"]:
            apply_excel_conditional_formatting(worksheet, column_letter, [0.1, 0.25], ["AFF2C9", "FFE699", "F7A9A9"])
        elif col_name in ["first-contentful-paint, sec"]:
            apply_excel_conditional_formatting(worksheet, column_letter, [1.8, 3], ["AFF2C9", "FFE699", "F7A9A9"])
        elif col_name in ["largest-contentful-paint, sec"]:
            apply_excel_conditional_formatting(worksheet, column_letter, [2.5, 4], ["AFF2C9", "FFE699", "F7A9A9"])

    # Apply all borders to the data cells
    apply_all_borders(worksheet)

    # Auto-adjust column widths
    auto_adjust_column_width(worksheet)

    writer.close()  # Correct method to finalize and save the file

except Exception as e:
    print(f"An error occurred: {e}")