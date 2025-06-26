import logging
import json
import traceback
import tempfile
import os
from datetime import datetime
from typing import Type
from langchain_core.tools import BaseTool, ToolException
from pydantic.fields import Field
from pydantic import create_model, BaseModel
from .api_wrapper import CarrierAPIWrapper


logger = logging.getLogger(__name__)


class CreateUIExcelReportTool(BaseTool):
    api_wrapper: CarrierAPIWrapper = Field(..., description="Carrier API Wrapper instance")
    name: str = "create_excel_report_ui"
    description: str = "Create Excel report from UI test results JSON files from the Carrier platform."
    args_schema: Type[BaseModel] = create_model(
        "CreateUIExcelReportInput",
        report_id=(str, Field(default="", description="UI Report ID to generate Excel report for")),
    )
    def _run(self, report_id: str = ""):
        try:
            # Check if report_id is provided
            if not report_id or report_id.strip() == "":
                return self._missing_input_response()
            
            # Get UI reports list and find the specific report
            ui_reports = self.api_wrapper.get_ui_reports_list()
            
            # Find the report by ID
            target_report = None
            for report in ui_reports:
                if str(report.get("id")) == str(report_id):
                    target_report = report
                    break
            
            if not target_report:
                return self._show_available_reports_message()
            
            return self._process_ui_report(target_report, report_id)
            
        except Exception:
            stacktrace = traceback.format_exc()
            logger.error(f"Error creating UI Excel report: {stacktrace}")
            raise ToolException(stacktrace)
    def _missing_input_response(self):
        """Response when report_id is missing."""
        return "Please provide me test id for generating excel report from your UI test"
    
    def _show_available_reports_message(self):
        """Show available reports when no matching report_id found."""
        try:
            ui_reports = self.api_wrapper.get_ui_reports_list()
            
            if not ui_reports:
                return "❌ **No UI test reports found.**"
            
            message = ["# ❌ No report found for the specified report ID\n"]
            message.append("## Available Report IDs:")
            
            for report in ui_reports[:10]:  # Show first 10 reports
                report_id = report.get("id")
                report_name = report.get("name", "Unnamed Report")
                test_status = report.get("test_status", "Unknown")
                start_time = report.get("start_time", "")
                
                message.append(f"- **Report ID: {report_id}** - {report_name} ({test_status}) - {start_time}")
            
            if len(ui_reports) > 10:
                message.append(f"... and {len(ui_reports) - 10} more reports")
            
            message.append("\n## 💡 Example:")
            message.append("```")
            message.append(f"report_id: {ui_reports[0].get('id') if ui_reports else 'YOUR_REPORT_ID'}")
            message.append("```")
            return "\n".join(message)
            
        except Exception:
            return "❌ **Error retrieving available report IDs. Please check your report_id and try again.**"
    
    def _process_ui_report(self, report, report_id):
        """Process single UI report and generate Excel file."""
        try:            # Get the UID from the report (similar to get_ui_report_by_id logic)
            uid = report.get("uid")
            if not uid:
                return f"❌ **No UID found for report {report_id}. Cannot process this report.**"
            
            # Create Excel reporter instance
            current_date = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            excel_file_name = f'/tmp/ui_report_{report_id}_{current_date}.xlsx'
            
            excel_reporter = LighthouseExcelReporter(excel_file_name)
            
            processed_files = 0
            
            # Get report links using the same method as GetUIReportByIDTool
            report_links = self.api_wrapper.get_ui_report_links(uid)
            
            if not report_links:
                return f"❌ **No report links found for report {report_id}.**"
            
            # Process each report link by converting HTML to JSON
            for html_url in report_links:
                try:
                    # Convert HTML URL to JSON URL by replacing .html with .json
                    json_url = html_url.replace('.html', '.json')
                    
                    # Extract file name from URL for worksheet naming
                    json_file_name = json_url.split('/')[-1]
                    
                    # Download JSON content directly from the converted URL
                    # Extract bucket and file name from the URL structure
                    # URL format: https://platform.getcarrier.io/api/v1/artifacts/artifact/default/{project_id}/reports/{file_name}
                    url_parts = json_url.split('/')
                    if len(url_parts) >= 2:
                        bucket = url_parts[-2]  # "reports"
                        file_name = url_parts[-1]  # actual file name
                    else:
                        bucket = "reports"
                        file_name = json_file_name
                    
                    json_content = self.api_wrapper.download_ui_report_json(bucket, file_name)
                    
                    if json_content:
                        # Create worksheet name from JSON file name
                        worksheet_name = self._create_worksheet_name(json_file_name)
                        
                        # Process JSON and add to Excel
                        excel_reporter.add_json_report(json_content, worksheet_name)
                        processed_files += 1
                        
                except Exception as e:
                    logger.error(f"Error processing JSON file from {html_url}: {e}")
                    continue
            
            if processed_files == 0:
                return f"❌ **No JSON files could be processed for report {report_id}.**"
            
            # Finalize Excel file
            excel_reporter.finalize()
            
            # Upload to Carrier artifacts
            report_name = report.get("name", f"report_{report_id}")
            bucket_name = report_name.replace("_", "").replace(" ", "").lower()
            excel_file_basename = os.path.basename(excel_file_name)
            
            self.api_wrapper.upload_file(bucket_name, excel_file_name)
            
            # Clean up temporary file
            if os.path.exists(excel_file_name):
                os.remove(excel_file_name)
            
            download_link = f"{self.api_wrapper.url.rstrip('/')}/api/v1/artifacts/artifact/default/{self.api_wrapper.project_id}/{bucket_name}/{excel_file_basename}"
            
            return f"""# ✅ UI Excel Report Generated Successfully!

## Report Information:
- **Report ID:** `{report_id}`
- **Report Name:** `{report.get("name", "N/A")}`
- **JSON Files Processed:** `{processed_files}`
- **Excel File:** `{excel_file_basename}`
- **Bucket:** `{bucket_name}`

## 📥 Download Link:
[Download Excel Report]({download_link})

## 🎯 What's included:
- Multiple worksheets for each JSON report file
- Lighthouse performance metrics formatted for analysis
- Conditional formatting for easy identification of performance issues"""
            
        except Exception as e:
            logger.error(f"Error processing UI report: {e}")
            raise ToolException(f"Error processing UI report: {e}")
        
    def _create_worksheet_name(self, json_file_name):
        """Create a valid worksheet name from JSON file name."""
        # Remove .json extension
        name = json_file_name.replace('.json', '')
        
        # Replace : with _ as specified in requirements
        name = name.replace(':', '_')
        
        # Excel worksheet names have limitations
        # Max 31 characters, no special characters except underscore
        name = name.replace('/', '_').replace('\\', '_').replace('[', '_').replace(']', '_')
        name = name.replace('*', '_').replace('?', '_').replace(':', '_')
        
        # Extract only the timestamp part (remove everything after the time)
        # Expected format: "24Jun2025_02_14_37_user-flow.re" -> "24Jun2025_02_14_37"
        parts = name.split('_')
        if len(parts) >= 4:
            # Keep first 4 parts which should be: date + 3 time parts
            # Example: ["24Jun2025", "02", "14", "37", "user-flow.re"] -> ["24Jun2025", "02", "14", "37"]
            timestamp_parts = parts[:4]
            name = '_'.join(timestamp_parts)
        
        # Ensure it's within Excel's 31 character limit
        if len(name) > 31:
            name = name[:31]
        
        return name


class LighthouseExcelReporter:
    """Excel reporter for Lighthouse UI test results."""
    
    def __init__(self, output_file):
        """Initialize the Excel reporter."""
        self.output_file = output_file
        self.workbook = None
        self.worksheets = {}
        
        # Import required libraries
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, PatternFill, Border, Side
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.utils import get_column_letter
            
            self.pd = pd
            self.Workbook = Workbook
            self.Alignment = Alignment
            self.PatternFill = PatternFill
            self.Border = Border
            self.Side = Side
            self.CellIsRule = CellIsRule
            self.get_column_letter = get_column_letter
            
            self.workbook = Workbook()
            # Remove default sheet
            if self.workbook.worksheets:
                self.workbook.remove(self.workbook.active)
                
        except ImportError as e:
            raise ToolException(f"Required libraries not available: {e}")
    
    def add_json_report(self, json_content, worksheet_name):
        """Add a JSON report as a new worksheet."""
        try:
            # Parse JSON content
            if isinstance(json_content, str):
                data = json.loads(json_content)
            else:
                data = json_content
            
            # Process Lighthouse data similar to the reference file
            data_rows = self._process_lighthouse_data(data)
            
            if not data_rows:
                logger.warning(f"No data extracted from JSON for worksheet {worksheet_name}")
                return
            
            # Create DataFrame
            df = self.pd.DataFrame(data_rows)
            
            if df.empty:
                logger.warning(f"Empty DataFrame for worksheet {worksheet_name}")
                return
            
            # Create pivot table
            df_pivot = df.pivot_table(index="Step name", columns="Audit", values="Numeric Value", aggfunc='mean')
            df_pivot = df_pivot.fillna('')
            
            # Add worksheet
            ws = self.workbook.create_sheet(title=worksheet_name)
            
            # Write data to worksheet
            self._write_dataframe_to_worksheet(df_pivot, ws)
            
            # Apply formatting
            self._apply_formatting(ws, df_pivot)
            
            logger.info(f"Added worksheet: {worksheet_name}")
            
        except Exception as e:
            logger.error(f"Error processing JSON report for worksheet {worksheet_name}: {e}")
    
    def _process_lighthouse_data(self, data):
        """Process Lighthouse JSON data similar to the reference implementation."""
        from urllib.parse import urlparse
        from collections import OrderedDict
        import re
        
        # Performance audits to extract (from reference file)
        performance_audits = [
            'first-contentful-paint',
            'speed-index',
            'interactive',
            'total-blocking-time',
            'largest-contentful-paint',
            'cumulative-layout-shift',
            'network-requests',
            'bootup-time',
            'interaction-to-next-paint',
            'server-response-time',
        ]
        
        # Audit naming mappings (from reference file)
        sec_audits = ['first-contentful-paint', 'interactive', 'largest-contentful-paint', 'mainthread-work-breakdown', 'network-requests', 'speed-index', 'javaScript-execution-time']
        ms_audits = ['interaction-to-next-paint', 'total-blocking-time', 'time-to-first-byte']
        
        rename_audits = {
            'bootup-time': 'javaScript-execution-time',
            'server-response-time': 'time-to-first-byte'
        }
        
        def extract_application_name(url):
            parsed_url = urlparse(url)
            hostname_parts = parsed_url.hostname.split('.') if parsed_url.hostname else []
            application_name = hostname_parts[0] if len(hostname_parts) > 1 else '3rd-party'
            return application_name
        
        data_rows = []
        step_order = OrderedDict()
        
        # Process steps from the data
        steps = data.get('steps', [])
        if not steps:
            # If no steps, treat the entire data as a single step
            steps = [{'name': 'main_report', 'lhr': data}]
        
        for index, step in enumerate(steps):
            step_name = step.get('name', f'step_{index}')
            step_order[step_name] = index
            lhr_data = step.get('lhr', {})
            
            url = lhr_data.get('finalDisplayedUrl', '3rd-party')
            application_name = extract_application_name(url)
            
            performance_score = lhr_data.get('categories', {}).get('performance', {}).get('score')
            performance_score = performance_score * 100 if performance_score is not None else None
            
            for audit in performance_audits:
                audit_result = lhr_data.get('audits', {}).get(audit, {})
                numeric_value = audit_result.get('displayValue')
                
                if numeric_value is not None:
                    numeric_value = re.sub(r'[a-zA-Z\s]', '', str(numeric_value))
                    if numeric_value:
                        numeric_value = numeric_value.replace(',', '')
                        try:
                            numeric_value = float(numeric_value)
                        except ValueError:
                            numeric_value = None
                    else:
                        numeric_value = None
                
                audit_display_name = rename_audits.get(audit, audit)
                if audit_display_name in sec_audits:
                    audit_display_name += ", sec"
                elif audit_display_name in ms_audits:
                    audit_display_name += ", ms"
                
                data_row = {
                    "Step name": step_name,
                    "Performance Score": performance_score,
                    "Audit": audit_display_name,
                    "Numeric Value": numeric_value
                }
                data_rows.append(data_row)
        
        return data_rows
    
    def _write_dataframe_to_worksheet(self, df, ws):
        """Write pandas DataFrame to Excel worksheet."""
        # Write headers
        ws.cell(row=1, column=1, value="Step name")
        for col_idx, col_name in enumerate(df.columns, 2):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, (index, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=index)
            for col_idx, value in enumerate(row, 2):
                ws.cell(row=row_idx, column=col_idx, value=value if value != '' else None)
    
    def _apply_formatting(self, ws, df):
        """Apply Excel formatting to the worksheet."""
        # Apply header formatting
        header_fill = self.PatternFill(start_color="7FD5D8", end_color="7FD5D8", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
        
        # Set alignment for 'Step name' column
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = self.Alignment(horizontal='left')
        
        # Apply conditional formatting
        for col_index, col_name in enumerate(df.columns, 2):
            column_letter = self.get_column_letter(col_index)
            
            if col_name in ["cumulative-layout-shift"]:
                self._apply_conditional_formatting(ws, column_letter, [0.1, 0.25], ["AFF2C9", "FFE699", "F7A9A9"])
            elif col_name in ["first-contentful-paint, sec"]:
                self._apply_conditional_formatting(ws, column_letter, [1.8, 3], ["AFF2C9", "FFE699", "F7A9A9"])
            elif col_name in ["largest-contentful-paint, sec"]:
                self._apply_conditional_formatting(ws, column_letter, [2.5, 4], ["AFF2C9", "FFE699", "F7A9A9"])
        
        # Apply borders
        self._apply_borders(ws)
        
        # Auto-adjust column widths
        self._auto_adjust_column_width(ws)
    
    def _apply_conditional_formatting(self, ws, column_letter, thresholds, colors):
        """Apply conditional formatting to a column."""
        ws.conditional_formatting.add(
            f'{column_letter}2:{column_letter}{ws.max_row}',
            self.CellIsRule(
                operator='lessThanOrEqual', 
                formula=[str(thresholds[0])], 
                stopIfTrue=True, 
                fill=self.PatternFill(start_color=colors[0], end_color=colors[0], fill_type="solid")
            )
        )
        ws.conditional_formatting.add(
            f'{column_letter}2:{column_letter}{ws.max_row}',
            self.CellIsRule(
                operator='between', 
                formula=[str(thresholds[0]+0.0001), str(thresholds[1])], 
                stopIfTrue=True, 
                fill=self.PatternFill(start_color=colors[1], end_color=colors[1], fill_type="solid")
            )
        )
        ws.conditional_formatting.add(
            f'{column_letter}2:{column_letter}{ws.max_row}',
            self.CellIsRule(
                operator='greaterThanOrEqual', 
                formula=[str(thresholds[1]+0.0001)], 
                stopIfTrue=True, 
                fill=self.PatternFill(start_color=colors[2], end_color=colors[2], fill_type="solid")
            )
        )
    
    def _apply_borders(self, ws):
        """Apply borders to all data cells."""
        thin_border = self.Border(
            left=self.Side(style='thin'), 
            right=self.Side(style='thin'), 
            top=self.Side(style='thin'), 
            bottom=self.Side(style='thin')
        )
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
    
    def _auto_adjust_column_width(self, ws):
        """Auto-adjust column widths."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    def finalize(self):
        """Finalize and save the Excel file."""
        if self.workbook and self.workbook.worksheets:
            self.workbook.save(self.output_file)
        else:
            raise ToolException("No worksheets were created")
