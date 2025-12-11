# Copyright (c) 2023 Artem Rozumenko
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
import io
import os
from typing import Iterator
import pandas as pd
from json import loads

from openpyxl import load_workbook
from xlrd import open_workbook
from langchain_core.documents import Document
from .AlitaTableLoader import AlitaTableLoader
from alita_sdk.runtime.langchain.constants import LOADER_MAX_TOKENS_DEFAULT

cell_delimiter = " | "

class AlitaExcelLoader(AlitaTableLoader):
    sheet_name: str = None
    file_name: str = None
    max_tokens: int = LOADER_MAX_TOKENS_DEFAULT
    add_header_to_chunks: bool = False
    header_row_number: int = 1

    def __init__(self, **kwargs):
        if not kwargs.get('file_path'):
            file_content = kwargs.get('file_content')
            if file_content:
                self.file_name = kwargs.get('file_name')
                kwargs['file_path'] = io.BytesIO(file_content)
        else:
            self.file_name = kwargs.get('file_path')
        super().__init__(**kwargs)
        self.sheet_name = kwargs.get('sheet_name')
        # Set and validate chunking parameters only once
        self.max_tokens = int(kwargs.get('max_tokens', LOADER_MAX_TOKENS_DEFAULT))
        self.add_header_to_chunks = bool(kwargs.get('add_header_to_chunks', False))
        header_row_number = kwargs.get('header_row_number', 1)
        # Validate header_row_number
        try:
            header_row_number = int(header_row_number)
            if header_row_number > 0:
                self.header_row_number = header_row_number
            else:
                self.header_row_number = 1
                self.add_header_to_chunks = False
        except (ValueError, TypeError):
            self.header_row_number = 1
            self.add_header_to_chunks = False

    def get_content(self):
        try:
            # Determine file extension
            file_extension = os.path.splitext(self.file_name)[-1].lower()

            if file_extension == '.xlsx':
                # Use openpyxl for .xlsx files
                return self._read_xlsx()
            elif file_extension == '.xls':
                # Use xlrd for .xls files
                return self._read_xls()
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
        except Exception as e:
            return f"Error reading Excel file: {e}"

    def _read_xlsx(self):
        """
        Reads .xlsx files using openpyxl.
        """
        workbook = load_workbook(self.file_path, data_only=True)  # `data_only=True` ensures we get cell values, not formulas
        sheets = workbook.sheetnames
        if self.sheet_name:
            if self.sheet_name in sheets:
                sheet_content = self.parse_sheet(workbook[self.sheet_name])
            else:
                sheet_content = [f"Sheet '{self.sheet_name}' does not exist in the workbook."]
            return {self.sheet_name: sheet_content}
        else:
            # Dictionary comprehension for all sheets
            return {name: self.parse_sheet(workbook[name]) for name in sheets}

    def _read_xls(self):
        """
        Reads .xls files using xlrd.
        """
        workbook = open_workbook(filename=self.file_name, file_contents=self.file_content)
        sheets = workbook.sheet_names()
        if self.sheet_name:
            if self.sheet_name in sheets:
                sheet = workbook.sheet_by_name(self.sheet_name)
                return {self.sheet_name: self.parse_sheet_xls(sheet)}
            else:
                return {self.sheet_name: [f"Sheet '{self.sheet_name}' does not exist in the workbook."]}
        else:
            # Dictionary comprehension for all sheets
            return {name: self.parse_sheet_xls(workbook.sheet_by_name(name)) for name in sheets}

    def parse_sheet(self, sheet):
        """
        Parses a single .xlsx sheet, extracting text and hyperlinks, and formats them.
        """
        sheet_content = []

        for row in sheet.iter_rows():
            row_content = []
            for cell in row:
                if cell.hyperlink:
                    # If the cell has a hyperlink, format it as Markdown
                    hyperlink = cell.hyperlink.target
                    cell_value = cell.value or ''  # Use cell value or empty string
                    row_content.append(f"[{cell_value}]({hyperlink})")
                else:
                    # If no hyperlink, use the cell value (computed value if formula)
                    row_content.append(str(cell.value) if cell.value is not None else "")
            # Join the row content into a single line using `|` as the delimiter
            sheet_content.append(cell_delimiter.join(row_content))

        # Format the sheet content based on the return type
        return self._format_sheet_content(sheet_content)

    def parse_sheet_xls(self, sheet):
        """
        Parses a single .xls sheet using xlrd, extracting text and hyperlinks, and formats them.
        """
        sheet_content = []

        # Extract hyperlink map (if available)
        hyperlink_map = getattr(sheet, 'hyperlink_map', {})

        for row_idx in range(sheet.nrows):
            row_content = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                cell_value = cell.value

                # Check if the cell has a hyperlink
                cell_address = (row_idx, col_idx)
                if cell_address in hyperlink_map:
                    hyperlink = hyperlink_map[cell_address].url_or_path
                    if cell_value:
                        row_content.append(f"[{cell_value}]({hyperlink})")
                else:
                    row_content.append(str(cell_value) if cell_value is not None else "")
            # Join the row content into a single line using `|` as the delimiter
            sheet_content.append(cell_delimiter.join(row_content))

        # Format the sheet content based on the return type
        return self._format_sheet_content(sheet_content)

    def _format_sheet_content(self, rows):
        """
        Specification:
        Formats a list of sheet rows into a list of string chunks according to the following rules:
        1. If max_tokens < 1, returns a single chunk (list of one string) with all rows joined by a newline ('\n').
           - If add_header_to_chunks is True and header_row_number is valid, the specified header row is prepended as the first line.
        2. If max_tokens >= 1:
           a. Each chunk is a string containing one or more rows, separated by newlines ('\n'), such that the total token count (as measured by tiktoken) does not exceed max_tokens.
           b. If add_header_to_chunks is True and header_row_number is valid, the specified header row is prepended once at the top of each chunk (not before every row).
           c. If a single row exceeds max_tokens, it is placed in its own chunk without splitting, with the header prepended if applicable.
        3. Returns: List[str], where each string is a chunk ready for further processing.
        """
        import tiktoken
        encoding = tiktoken.get_encoding('cl100k_base')

        # --- Inner functions ---
        def count_tokens(text):
            """Count tokens in text using tiktoken encoding."""
            return len(encoding.encode(text))

        def finalize_chunk(chunk_rows):
            """Join rows for a chunk, prepending header if needed."""
            if self.add_header_to_chunks and header:
                return '\n'.join([header] + chunk_rows)
            else:
                return '\n'.join(chunk_rows)
        # --- End inner functions ---

        # If max_tokens < 1, return all rows as a single chunk
        if self.max_tokens < 1:
            return ['\n'.join(rows)]

        # Extract header if needed
        header = None
        if self.add_header_to_chunks and rows:
            header_idx = self.header_row_number - 1
            header = rows.pop(header_idx)

        chunks = []  # List to store final chunks
        current_chunk = []  # Accumulate rows for the current chunk
        current_tokens = 0  # Token count for the current chunk

        for row in rows:
            row_tokens = count_tokens(row)
            # If row itself exceeds max_tokens, flush current chunk and add row as its own chunk (with header if needed)
            if row_tokens > self.max_tokens:
                if current_chunk:
                    chunks.append(finalize_chunk(current_chunk))
                    current_chunk = []
                    current_tokens = 0
                # Add the large row as its own chunk, with header if needed
                if self.add_header_to_chunks and header:
                    chunks.append(finalize_chunk([row]))
                else:
                    chunks.append(row)
                continue
            # If adding row would exceed max_tokens, flush current chunk and start new
            if current_tokens + row_tokens > self.max_tokens:
                if current_chunk:
                    chunks.append(finalize_chunk(current_chunk))
                current_chunk = [row]
                current_tokens = row_tokens
            else:
                current_chunk.append(row)
                current_tokens += row_tokens
        # Add any remaining rows as the last chunk
        if current_chunk:
            chunks.append(finalize_chunk(current_chunk))
        return chunks

    def load(self) -> list:
        docs = []
        content_per_sheet = self.get_content()
        # content_per_sheet is a dict of sheet_name: list of chunk strings
        for sheet_name, content_chunks in content_per_sheet.items():
            metadata = {
                "source": f'{self.file_path}:{sheet_name}',
                "sheet_name": sheet_name,
                "file_type": "excel",
            }
            # Each chunk is a separate Document
            for chunk in content_chunks:
                docs.append(Document(page_content=chunk, metadata=metadata))
        return docs

    def read(self, lazy: bool = False):
        df = pd.read_excel(self.file_path, sheet_name=None, engine='calamine')
        docs = []
        for key in df.keys():
            if self.raw_content:
                docs.append(df[key].to_string())
            else:
                for record in loads(df[key].to_json(orient='records')):
                    docs.append(record)
        return docs

    def read_lazy(self) -> Iterator[dict]:
        df = pd.read_excel(self.file_path, sheet_name=None, engine='calamine')
        for key in df.keys():
            if self.raw_content:
                yield df[key].to_string()
            else:
                for record in loads(df[key].to_json(orient='records')):
                    yield record
        return
